from tkinter import *
from tkinter import filedialog
import xlrd
import xlsxwriter
import xlwings
import xlwt
import openpyxl as px
from os import *

print(u'検索文字を入力してください')
word = input('>>')

sheet_name_array = list
array_list = [word]

select_file = Tk()
select_file.withdraw()
full_path = filedialog.askopenfilenames()

row_buf = ''
for f in full_path:
    drive, tail = path.split(f)
    root, ext = path.splitext(f)
    wb = xlrd.open_workbook(f)
    sheet_name_array = wb.sheet_names()
    for sheet_name in sheet_name_array:
        sheet = wb.sheet_by_name(sheet_name)
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                for item in array_list:
                    if item in str(sheet.cell(row, col).value) and not row == row_buf:
                        row_buf = row
                        print(sheet.cell(row, col), tail)

                        """
                        print(sheet.cell(row, 0).value, sheet.cell(row, 3).value)
                        """
