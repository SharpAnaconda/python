import tkinter as tk
import xlrd
import xlsxwriter
import xlwings
import xlwt
import openpyxl as px
from tkinter import filedialog
from os import path
import datetime

select_file = tk.Tk()
select_file.withdraw()

wb2 = xlrd.open_workbook(filedialog.askopenfilename())
sheet2 = wb2.sheet_by_name(u'縁成')
full_path = filedialog.askopenfilenames()

for f in full_path:
    drive, tail = path.split(f)
    root, ext = path.splitext(f)
    print(tail)
    wb = px.load_workbook(f)
    sheet1 = wb[u'許可願い書']
    for row in sheet1.iter_rows():
        for cell in row:
            if cell.value == u'発行№':
                value: str = sheet1[str(cell.column + str(int(cell.row) + 2))].value

    for row in range(sheet2.nrows):
        for col in range(sheet2.ncols):
            if str(sheet2.cell(row, col).value) == value:
                value2: str = sheet2.cell(row, col - 1).value

    for row in sheet1.iter_rows():
        for cell in row:
            if str(cell.value)[0:4] == u'管理番号':
                cell.value = u'管理番号: ' + value2
            if str(cell.value)[0:2] == u'日付':
                cell.value = u'日付: ' + str(datetime.date.today())

    a, b = path.split(f)
    temp: str = a + "/新しいフォルダー (4)/" + b
    wb.save(temp)
