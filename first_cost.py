import tkinter as tk
import xlrd
import xlsxwriter
import xlwings
import xlwt
import openpyxl as px
from tkinter import filedialog
from os import path

select_file = tk.Tk()
select_file.withdraw()

wb_2 = px.Workbook()
ws_2 = wb_2[u'Sheet']

i = 0

full_path = filedialog.askopenfilenames()
for f in full_path:
    drive, tail = path.split(f)
    root, ext = path.splitext(f)
    wb = xlrd.open_workbook(f)
    sheet_1 = wb.sheet_by_name(u'許可願い書')
    array = []
    temp = []
    for row in range(sheet_1.nrows):
        for col in range(sheet_1.ncols):
            temp.append(sheet_1.cell(row, col).value)
            string = str(sheet_1.cell(row, col).value)
            if string[0:4] == u'管理番号':
                kannribanngou_address = [row, col]
            elif string[0:2] == u'日付':
                hiduke_address = [row, col]
            elif string[0:3] == u'発行№':
                hakkou_address = [row, col]
            elif string[0:3] == u'部品名':
                buhinnmei_address = [row, col]
        array.append(temp)
        temp = []

    for row in range(hakkou_address[0] + 2, sheet_1.nrows):
        i += 1
        if array[row][buhinnmei_address[1]] == "":
            i -= 1
        else:
            for col in range(hakkou_address[1], len(array[row])):
                ws_2.cell(row=i, column=2 + col - hakkou_address[1]).value = array[row][col]
                print(array[row][col])
                if col == 1:
                    ws_2.cell(row=i, column=1).value = " "
                if (col == 1 or col == 2) and array[row][col] == "":
                    array[row][col] = array[row - 1][col]
                    ws_2.cell(row=i, column=2 + col - hakkou_address[1]).value = array[row][col]
                    print(array[row][col])
    wb_2.save("C:\\Users\\s138093\\Desktop\\sample.xlsx")
