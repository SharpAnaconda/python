import tkinter as tk
import xlrd
import xlsxwriter
import xlwings
import xlwt
import openpyxl as px
from tkinter import filedialog
from os import path

empty_set = set()
empty_list = list()
aggr_list = list()

cnt = 0
temp = 0

sheet_name_array = []
hoge = []

select_file = tk.Tk()
select_file.withdraw()
full_path = filedialog.askopenfilenames()

for f in full_path:
    drive, tail = path.split(f)
    root, ext = path.splitext(f)
    wb = xlrd.open_workbook(f)
    sheet_name_array = wb.sheet_names()
    sheet = wb.sheet_by_name('解析結果')
    for row in range(3, sheet.nrows):
        hoge = str(sheet.cell(row, 3).value).split('\n')
        for hogehoge in hoge:
            empty_set.add(hogehoge)

empty_set.remove('')
empty_set.remove('42')

empty_list.extend(empty_set)
empty_list.sort()

for f in full_path:
    wb = xlrd.open_workbook(f)
    sheet_name_array = wb.sheet_names()
    sheet = wb.sheet_by_name('解析結果')
    for lot in empty_list:
        for row in range(3,sheet.nrows):
            hoge = str(sheet.cell(row, 3).value).split('\n')
            if lot in str(sheet.cell(row, 3).value):
                if sheet.cell(row, 18).value != '':
                    cnt += 1
                    if sheet.cell(row, 18).value != '——':
                        temp = temp + sheet.cell(row, 18).value / len(hoge)
        if cnt != 0:
            aggr_list.append(temp / cnt)
            temp = 0
            cnt = 0
        else:
            aggr_list.append(temp)
            temp = 0
            cnt = 0
print(empty_list)
print(aggr_list)